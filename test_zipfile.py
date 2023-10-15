from utils import RESOURCES_PATH, TMP_PATH
from zipfile import ZipFile
import os
import openpyxl
from pypdf import PdfReader

def test_zipfile():
    if not os.path.exists(TMP_PATH):
        os.mkdir(TMP_PATH)

    with ZipFile(f'{TMP_PATH}/zipfile.zip', 'w') as zf:
        for root, dirs, files in os.walk(RESOURCES_PATH):
            for file in files:
                add_file = os.path.join(root, file)
                file_path = os.path.relpath(add_file, RESOURCES_PATH)
                zf.write(add_file, file_path)

    with ZipFile(f'{TMP_PATH}/zipfile.zip', mode='r') as zf:
        for file in zf.namelist():
            print(file)

    with ZipFile(f'{TMP_PATH}/zipfile.zip') as zf:
         for file in zf.namelist():
            if file == os.path.basename(os.path.join(RESOURCES_PATH, 'file_example_XLS_10.xls')):
                print(f'Файл {file} в архиве zipfile.zip соответствует файлу из каталога {RESOURCES_PATH}')
            if file == os.path.basename(os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx')):
                print(f'Файл {file} в архиве zipfile.zip соответствует файлу из каталога {RESOURCES_PATH}')
            if file == os.path.basename(os.path.join(RESOURCES_PATH, 'Hello.txt')):
                print(f'Файл {file} в архиве zipfile.zip соответствует файлу из каталога {RESOURCES_PATH}')
            if file == os.path.basename(os.path.join(RESOURCES_PATH, 'Python Testing with Pytest (Brian Okken).pdf')):
                print(f'Файл {file} в архиве zipfile.zip соответствует файлу из каталога {RESOURCES_PATH}')

def test_pdf():
    with ZipFile(f'{TMP_PATH}/zipfile.zip') as zf:
        pdf_file = PdfReader(zf.open('Python Testing with Pytest (Brian Okken).pdf', 'r'))
        len_pdf_zip = len(pdf_file.pages)
        text_pdf_zip = pdf_file.pages[1].extract_text()

    reader = PdfReader(os.path.join(RESOURCES_PATH, "Python Testing with Pytest (Brian Okken).pdf"))
    len_pdf_res = len(reader.pages)
    text_pdf_res = reader.pages[1].extract_text()

    assert len_pdf_zip == len_pdf_res
    assert text_pdf_zip == text_pdf_res

def test_txt():
    with ZipFile(f'{TMP_PATH}/zipfile.zip') as zf:
        text = zf.read('Hello.txt').decode(encoding="utf-8")

    with open(os.path.abspath('resources/Hello.txt'), 'r') as f:
        text_res = f.read()

    assert text == text_res

def test_xls():
    with ZipFile(f'{TMP_PATH}/zipfile.zip') as zf:
        xls = zf.getinfo('file_example_XLS_10.xls')

        assert xls.file_size == os.path.getsize('resources/file_example_XLS_10.xls')


def test_xlsx():
    with ZipFile(f'{TMP_PATH}/zipfile.zip') as zf:
        zip_book = openpyxl.load_workbook(zf.open('file_example_XLSX_50.xlsx', 'r'))
        zip_sheet = zip_book.active

    rec_book = openpyxl.load_workbook(os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx'))
    rec_sheet = rec_book.active

    assert zip_sheet.cell(row=3, column=2).value == rec_sheet.cell(row=3, column=2).value
